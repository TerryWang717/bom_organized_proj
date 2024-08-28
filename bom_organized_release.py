#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@FileName     :bom_organized_release.py
@Function     :本程序是根据凡总需求, 实现Joyrock标准BOM表的整理。
               将xlsx格式的BOM文件以"元件值", "封装" 和 "精度" 这三列为参考依据进行操作,
               这三列都相同的视为同一元器件，将同一元器件的位号合并到一行中, 并重新统计数量。
@Description  :运行程序,弹出文件选择对话框, 选择需要整理的xlsx格式的BOM表, 整理好后会在原目录生成一个“原文件名+时间戳”的新文件。
@Date         :2024/08/27 
@Author       :Terry.Wang
@Email        :2948185878@qq.com
@Reference    :主要以徐晓康的整理脚本(https://blog.csdn.net/weixin_42837669/article/details/132052784)为基础,
               使用了pandas进行数据分析处理, openpyxl来调整Excel格式。
'''


import openpyxl
import pyautogui
import re
import sys
import os
import time
import pyperclip
import win32com.client as win32
import pandas as pd
import tkinter as tk
from tkinter import ttk,messagebox,filedialog

# 版本标识
Major_Version = "1"
Minor_Version = "0"
Revision_Version = "1"
Suffix = "release"

__version__ = Major_Version + '.' + Minor_Version + '.' + Revision_Version + '_' + Suffix
print(f"当前版本：{__version__}")

# 主窗口实例化
root = tk.Tk()
root.title('文件路径选择')
max_w, max_h = root.maxsize()
root.geometry(f'500x300+{int((max_w - 500) / 2)}+{int((max_h - 300) / 2)}')  # 居中显示
root.resizable(width=False, height=False)
# 隐藏主窗口，直接显示文件选择窗口
root.withdraw()

# 文件路径保存到bom_path中
bom_path = filedialog.askopenfilename(title='请选择要处理的文件')

# 表头
std_bom_header = ('No.', '元件类型', '物料编码', '元件值', '物料描述', '封装', '位号', '数量', '单价 RMB', '金额', '备注', 'Option', '精度', '耐压')


# %%
# 获取当前时间戳
timestamp = time.strftime("%Y%m%d%H%M%S", time.localtime())
print(f"当前时间戳: {timestamp}")


# %%
# 实际只用到了bom name
def get_sheet_and_bom_name(old_bom_excel_path):
    bom_to_excel_folder_path = os.path.dirname(old_bom_excel_path)
    bom_to_excel_filename = os.path.splitext(os.path.basename(old_bom_excel_path))[0]

    # 最终的sheet名, 长度不能超过31字符，否则excel打开会报警告
    final_sheet_title = (bom_to_excel_filename + '-' + timestamp)[:31]
    # 新的表格只能以.xlsx为后缀
    final_bom_name = bom_to_excel_folder_path + '\\' + bom_to_excel_filename + '-' + timestamp + '.xlsx'
    pandas_bom_name =  bom_to_excel_folder_path + '\\' + bom_to_excel_filename + '-' + 'pandas_bom' + '.xlsx'

    return final_sheet_title, final_bom_name, pandas_bom_name

final_sheet_title, final_bom_name, pandas_bom_name = get_sheet_and_bom_name(bom_path)

print(f'最终excel文件路径: {final_bom_name}')


# 复制文件路径到剪贴板，以便在下一个程序中粘贴
pyperclip.copy(final_bom_name)


# 打开Excel文件
workbook = openpyxl.load_workbook(bom_path)
origin_sheet = workbook.active

#添加“精度”和“耐压”标签
origin_sheet.cell(row=4, column=13, value='精度')
origin_sheet.cell(row=4, column=14, value='耐压')


# 暂时保存并关闭Excel文件
workbook.save(bom_path)
workbook.close()


############ pandas DataFrame analysis ###################

# pandas读取Excel文件，选择第4行（从0开始算）为列名
df = pd.read_excel(bom_path, header=3)


# 根据'NC'或'DNP'字符串筛选，之后取反来删除这些行并重置索引
mask = ~((df['元件值'].str.startswith('NC')) |(df['元件值'].str.startswith('DNP'))| (df['Option'].str.startswith('DNP')))
df_del = df[mask].reset_index(drop=True)

### 分组聚合
# groupby函数用于分组，三列一样则去重。
# agg（）函数表示处理groupbuy处理的数据之外的列数据。并进行聚合
# reset_index()函数用来重置索引
df_new = df_del.groupby(['元件值', '封装', '精度'],sort=False, dropna=False).agg({'位号': lambda x:','.join(x.values)}).reset_index()

# 统计数量并新增到df_new 表格中
quantity = df_new['位号'].str.count(',') + 1
df_new['数量'] = quantity


# 在df_del删除NC后的表格上进行操作，删除重复行只保留第一个
df_drop = df_del.drop_duplicates(subset=['元件值', '封装', '精度'], ignore_index=True)


# 把之前统计好的位号和数量复制到df_drop表格中并保存为Excel文件
df_drop_cp = df_drop.copy()
df_drop_cp['位号'] = df_new['位号'] 
df_drop_cp['数量'] = df_new['数量']
df_drop = df_drop_cp.copy()

df_drop.to_excel(pandas_bom_name, index=False)


############ 打开Excel文件，现在切换为openpyxl来调整格式 ###################
workbook = openpyxl.load_workbook(pandas_bom_name)
output_sheet = workbook.active


# 填充No.列， 重排序
max_row = output_sheet.max_row 
for i in range(2, max_row + 1):
    output_sheet.cell(row=i, column=1).value = i - 1

# 设置精度为百分比格式
for cell in  output_sheet['M']:            
        cell.number_format = '0%'        

# 插入3个空行在表头，复制原表格的前3行到新表格
output_sheet.insert_rows(1,3)
origin_wb = openpyxl.load_workbook(bom_path)
origin_sheet = origin_wb.active
copy_rows = origin_sheet[1:3]
for row in copy_rows:
    for cell in row:
        output_sheet[cell.coordinate].value = cell.value


### 合并单元格
# 前3行合并单元格
output_sheet.merge_cells('A1:N1')
output_sheet.merge_cells('A2:D2')
output_sheet.merge_cells('A3:D3')

# B列根据字符串合并单元格

# 从第6行开始
previous_row = 6
for cell in output_sheet['B']:
    if cell.data_type == 's':     
        current_row = cell.row    
        if (current_row > 6) :            
            start_num = 'B' + str(previous_row)           
            end_num = 'B' + str(current_row - 1)
            previous_row = current_row        
            output_sheet.merge_cells(start_num + ':' + end_num)
    #最后一行为空值时的操作        
    if (cell.row ==  output_sheet.max_row) and (cell.value is None):        
        start_num = 'B' + str(previous_row)
        end_num = 'B' + str(cell.row)        
        output_sheet.merge_cells(start_num + ':' + end_num)        
                    


### 设置全局格式
# 设置字体样式和大小
font = openpyxl.styles.Font(name='Consolas', size=11)

# 设置文本对齐和自动换行
alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

# 遍历单元格，并应用格式，并给所有单元格加上边框
for row in output_sheet.iter_rows(max_col=len(std_bom_header)):
    for cell in row:
        #cell.font = font
        cell.alignment = alignment
        cell.border = openpyxl.styles.Border(
            left   = openpyxl.styles.Side (style = 'thin'),
            right  = openpyxl.styles.Side (style = 'thin'),
            top    = openpyxl.styles.Side (style = 'thin'),
            bottom = openpyxl.styles.Side (style = 'thin')
        )

# 设置列宽 依据std_bom_header = ('No.', '元件类型', '物料编码', '元件值', '物料描述', '封装', '位号', '数量', '单价 RMB', '金额', '备注', 'Option', '精度', '耐压')
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("No."         )+ 1)].width = 6
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("元件类型"     )+ 1)].width = 15
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("物料编码"         )+ 1)].width = 15
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("元件值"    )+ 1)].width = 20
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("物料描述"          )+ 1)].width = 20
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("封装"          )+ 1)].width = 20
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("位号"          )+ 1)].width = 50
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("数量"          )+ 1)].width = 10
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("单价 RMB"          )+ 1)].width = 6
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("金额"          )+ 1)].width = 6
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("备注"          )+ 1)].width = 6
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("Option"          )+ 1)].width = 8
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("精度"          )+ 1)].width = 8
output_sheet.column_dimensions[openpyxl.utils.get_column_letter (std_bom_header.index ("耐压"          )+ 1)].width = 8


### 特定行列设置格式
# 隐藏C列
output_sheet.column_dimensions['C'].hidden = True

# 第一行字体增大加粗，设置行高
font_1 = openpyxl.styles.Font(size=18, bold=True) 
for cell in output_sheet[1]:
    cell.font = font_1
output_sheet.row_dimensions[1].height = 50

# 第二行和第二列字体加粗
font_2 = openpyxl.styles.Font(bold=True) 
for cell in output_sheet[2]:
    cell.font = font_2

for cell in output_sheet['B']:
    cell.font = font_2     
  

# 根据“数量”列计算行数调整行高。
for cell in output_sheet['H']:
    if (cell.row > 4):
        # 向上取整
        lines = (cell.value // 12) + 1            
        output_sheet.row_dimensions[cell.row].height = (lines) * 14
     

# 保存最终文件
workbook.save(final_bom_name)
workbook.close()
# 删除中间文件
os.remove(pandas_bom_name)
pyautogui.alert(f'新BOM\n{final_bom_name}\n创建成功!', '确认')        
